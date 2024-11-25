import pandas as pd
import pyqtgraph as pg

from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QVBoxLayout, QPushButton, QLabel, QWidget, QSplitter, QTableView, QCalendarWidget, \
    QHBoxLayout, QDialog, QLCDNumber, QDateTimeEdit, QMessageBox, QFileDialog, QCheckBox, QLayout, QDateEdit
from PyQt5.QtGui import QIcon, QStandardItemModel, QStandardItem
from PyQt5.QtCore import Qt, QSortFilterProxyModel, QDateTime, QTime, QTimer, QDate

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import LineChart, Reference, BarChart

from pyqtgraph import DateAxisItem

from sqlalchemy import desc, column
from models.Report import SDM120Report, SDM630Report, SDM120ReportTmp


class DeviceDetailsSDM120Widget(QWidget):
    def __init__(self, main_window, device):
        super().__init__(main_window)
        self.device = device
        self.db_session = main_window.db_session
        self.main_window = main_window

        layout = QVBoxLayout(self)

        main_splitter = QSplitter(Qt.Horizontal)
        main_splitter.setChildrenCollapsible(False)
        layout.addWidget(main_splitter)

        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)

        table_widget = QWidget()
        table_layout = QVBoxLayout(table_widget)
        self.label = QLabel(f"Ім'я пристрою: {device.name} | Модель пристрою: {device.model}")
        table_layout.addWidget(self.label)

        button_layout = QHBoxLayout()

        refresh_button = QPushButton()
        refresh_button.setIcon(QIcon("pyqt/icons/refresh.png"))
        refresh_button.setFixedSize(36, 36)
        refresh_button.clicked.connect(self.load_report_data)
        button_layout.addWidget(refresh_button)

        export_button = QPushButton("Експорт в Excel")
        export_button.setFixedHeight(36)
        export_button.clicked.connect(self.open_export_dialog)
        button_layout.addWidget(export_button)

        table_layout.addLayout(button_layout)

        self.report_table = QTableView(self)
        table_layout.addWidget(self.report_table)
        left_layout.addWidget(table_widget)
        main_splitter.addWidget(left_widget)

        right_splitter = QSplitter(Qt.Vertical)
        right_splitter.setChildrenCollapsible(False)
        main_splitter.addWidget(right_splitter)

        top_right_widget = QWidget()
        top_right_layout = QVBoxLayout(top_right_widget)

        self.select_range_button = QPushButton("Обрати проміжок")
        self.select_range_button.clicked.connect(self.open_date_range_dialog)
        top_right_layout.addWidget(self.select_range_button)

        self.voltage_graph_widget = pg.PlotWidget()
        self.current_graph_widget = pg.PlotWidget()
        self.energy_graph_widget = pg.PlotWidget()

        top_right_layout.addWidget(self.voltage_graph_widget)
        top_right_layout.addWidget(self.current_graph_widget)
        top_right_layout.addWidget(self.energy_graph_widget)

        right_splitter.addWidget(top_right_widget)

        bottom_right_widget = QWidget()
        bottom_right_layout = QVBoxLayout(bottom_right_widget)

        # Годинник
        self.clock_label = QLabel()
        self.clock_label.setStyleSheet("font-size: 24pt;")
        bottom_right_layout.addWidget(self.clock_label)  # Додаємо годинник на верх

        # Блок з індикаторами в горизонтальному порядку
        indicators_layout = QHBoxLayout()

        # Індикатори
        self.voltage_lcd = QLCDNumber()
        self.voltage_lcd.setSegmentStyle(QLCDNumber.Flat)
        self.current_lcd = QLCDNumber()
        self.current_lcd.setSegmentStyle(QLCDNumber.Flat)
        self.energy_lcd = QLCDNumber()
        self.energy_lcd.setSegmentStyle(QLCDNumber.Flat)

        voltage_label = QLabel("Напруга (V)")
        voltage_label.setStyleSheet("font-size: 14pt; font-weight: bold;")
        current_label = QLabel("Струм (A)")
        current_label.setStyleSheet("font-size: 14pt; font-weight: bold;")
        energy_label = QLabel("Енергія (kWh)")
        energy_label.setStyleSheet("font-size: 14pt; font-weight: bold;")

        # Додаємо кожен індикатор у горизонтальний лейаут
        voltage_layout = QVBoxLayout()
        voltage_layout.addWidget(voltage_label)
        voltage_layout.addWidget(self.voltage_lcd)
        indicators_layout.addLayout(voltage_layout)

        current_layout = QVBoxLayout()
        current_layout.addWidget(current_label)
        current_layout.addWidget(self.current_lcd)
        indicators_layout.addLayout(current_layout)

        energy_layout = QVBoxLayout()
        energy_layout.addWidget(energy_label)
        energy_layout.addWidget(self.energy_lcd)
        indicators_layout.addLayout(energy_layout)

        # Додаємо горизонтальний блок індикаторів під годинником
        bottom_right_layout.addLayout(indicators_layout)

        right_splitter.addWidget(bottom_right_widget)

        self.set_splitter_fixed_ratios(main_splitter, [1, 1])
        self.set_splitter_fixed_ratios(right_splitter, [3, 1])

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_ui)
        self.timer.setInterval(1000)
        self.timer.start()

        self.load_report_data()
        self.plot_graphs()
        self.set_light_theme()

    def set_splitter_fixed_ratios(self, splitter, ratios):
        total = sum(ratios)
        sizes = [int(ratio / total * splitter.size().width()) for ratio in ratios]
        splitter.setSizes(sizes)
        splitter.handle(1).setEnabled(False)

    def update_ui(self):
        # Оновлення годинника
        self.update_clock()

        # Оновлення показників
        self.update_indicators()

    def update_indicators(self):
        last_report = self.db_session.query(SDM120ReportTmp).filter_by(device_id=self.device.id).order_by(desc(SDM120ReportTmp.timestamp)).first()
        if last_report:
            self.voltage_lcd.display(getattr(last_report, 'voltage', 0))
            self.current_lcd.display(getattr(last_report, 'current', 0))
            self.energy_lcd.display(getattr(last_report, 'total_active_energy', 0))

    def update_clock(self):
        # Оновлення поточного часу
        current_time = QTime.currentTime().toString("HH:mm:ss")
        self.clock_label.setText(f"Поточний час :{current_time}")

    def create_table_model(self, report_data, device):
        columns = set()
        parameter_units = {}
        if device.parameters:
            for param in device.parameters.split(','):
                param_name, unit = param.split(':')
                parameter_units[param_name.strip()] = unit.strip()

        for report in report_data:
            for column_name in report.__table__.columns.keys():
                if column_name not in ['id', 'device_id'] and getattr(report, column_name) is not None:
                    columns.add(column_name)

        columns = sorted(columns)
        model = QStandardItemModel(len(report_data), len(columns))
        header_labels = []
        for column in columns:
            unit = parameter_units.get(column, "")
            header_labels.append(f"{column} ({unit})" if unit else column)

        model.setHorizontalHeaderLabels(header_labels)

        for row, report in enumerate(report_data):
            for col, column in enumerate(columns):
                value = getattr(report, column)
                model.setItem(row, col, QStandardItem(str(value) if value is not None else ""))

        return model

    def load_report_data(self):
        if self.device.model == "SDM120":
            report_data = self.db_session.query(SDM120Report).filter_by(device_id=self.device.id).order_by(
                desc(SDM120Report.timestamp)).all()
        elif self.device.model == "SDM630":
            report_data = self.db_session.query(SDM630Report).filter_by(device_id=self.device.id).order_by(
                desc(SDM630Report.timestamp)).all()
        else:
            report_data = []

        model = self.create_table_model(report_data, self.device)

        proxy_model = QSortFilterProxyModel()
        proxy_model.setSourceModel(model)
        proxy_model.setSortCaseSensitivity(Qt.CaseInsensitive)

        self.report_table.setModel(proxy_model)

        self.report_table.sortByColumn(3, Qt.DescendingOrder)

        self.report_table.setSortingEnabled(True)
        self.report_table.resizeColumnsToContents()

        header = self.report_table.horizontalHeader()
        header.setSectionResizeMode(QtWidgets.QHeaderView.Fixed)

        self.update_indicators()

    def plot_voltage(self, timestamps, voltages):
        timestamps_numeric = [ts.timestamp() for ts in timestamps]

        self.voltage_graph_widget.clear()

        self.voltage_graph_widget.setAxisItems({'bottom': DateAxisItem()})

        self.voltage_graph_widget.plot(timestamps_numeric, voltages, pen=pg.mkPen(color='b', width=2), name="Напруга")
        self.voltage_graph_widget.setYRange(0, 400)  # Встановлюємо фіксований діапазон для напруги
        self.voltage_graph_widget.setLabel('bottom', 'Час')

    def plot_current(self, timestamps, currents):
        timestamps_numeric = [ts.timestamp() for ts in timestamps]

        self.current_graph_widget.clear()

        self.current_graph_widget.setAxisItems({'bottom': DateAxisItem()})

        self.current_graph_widget.plot(timestamps_numeric, currents, pen=pg.mkPen(color='r', width=2), name="Струм")
        self.current_graph_widget.setYRange(0, 200)  # Встановлюємо фіксований діапазон для струму
        self.current_graph_widget.setLabel('bottom', 'Час')

    def plot_energy(self, hourly_timestamps, hourly_energy):
        hourly_timestamps_numeric = [ts.timestamp() for ts in hourly_timestamps]

        bar_x = []
        bar_heights = []

        for i in range(len(hourly_timestamps_numeric)):
            start_time = hourly_timestamps_numeric[i]
            end_time = hourly_timestamps_numeric[i] + 3600  # 3600 секунд = 1 година

            bar_x.append((start_time, end_time))
            bar_heights.append(hourly_energy[i])

        self.energy_graph_widget.clear()

        self.energy_graph_widget.setAxisItems({'bottom': DateAxisItem()})

        bar_item = pg.BarGraphItem(x=[x[0] for x in bar_x], height=bar_heights, width=1800, brush='g')

        self.energy_graph_widget.addItem(bar_item)
        self.energy_graph_widget.setYRange(0, max(hourly_energy))
        self.energy_graph_widget.setLabel('bottom', 'Час')

    def plot_graphs(self, start_date=None, end_date=None):
        # Якщо дати не вказані, використовуємо повний діапазон
        query = self.db_session.query(SDM120Report).filter_by(device_id=self.device.id)
        if start_date and end_date:
            query = query.filter(SDM120Report.timestamp >= start_date, SDM120Report.timestamp <= end_date)
        report_data = query.order_by(desc(SDM120Report.timestamp)).all()

        if not report_data:
            QMessageBox.warning(self, "Попередження", "Дані відсутні для заданого періоду.")
            return

        timestamps = []
        voltages = []
        currents = []
        energies = []

        for report in report_data:
            timestamps.append(report.timestamp)
            voltages.append(report.voltage)
            currents.append(report.current)
            energies.append(report.total_active_energy)

        self.plot_voltage(timestamps, voltages)
        self.plot_current(timestamps, currents)

        # Підрахунок годинної енергії
        hourly_energy = []
        hourly_timestamps = []

        last_energy = None
        current_hour_start = None
        current_hour_energy = 0.0

        for report in report_data:
            current_hour = report.timestamp.replace(minute=0, second=0, microsecond=0)

            if current_hour_start is None:
                current_hour_start = current_hour

            if current_hour != current_hour_start:
                if last_energy is not None:
                    hourly_energy.append(current_hour_energy)
                    hourly_timestamps.append(current_hour_start)
                current_hour_start = current_hour
                current_hour_energy = 0.0

            if last_energy is not None:
                current_hour_energy += abs(report.total_active_energy - last_energy.total_active_energy)

            last_energy = report

        if last_energy is not None:
            hourly_energy.append(current_hour_energy)
            hourly_timestamps.append(current_hour_start)

        self.plot_energy(hourly_timestamps, hourly_energy)

    def open_date_range_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Обрати проміжок")
        dialog.setFixedSize(500, 500)

        layout = QVBoxLayout(dialog)

        start_calendar = QCalendarWidget()
        start_calendar.setGridVisible(True)
        layout.addWidget(QLabel("Початкова дата:"))
        layout.addWidget(start_calendar)

        end_calendar = QCalendarWidget()
        end_calendar.setGridVisible(True)
        layout.addWidget(QLabel("Кінцева дата:"))
        layout.addWidget(end_calendar)

        button_layout = QHBoxLayout()
        confirm_button = QPushButton("Підтвердити")
        cancel_button = QPushButton("Скасувати")
        button_layout.addWidget(confirm_button)
        button_layout.addWidget(cancel_button)

        layout.addLayout(button_layout)

        def apply_date_range():
            start_date = start_calendar.selectedDate().startOfDay().toPyDateTime()
            end_date = end_calendar.selectedDate().endOfDay().toPyDateTime()
            self.plot_graphs(start_date=start_date, end_date=end_date)
            dialog.accept()

        confirm_button.clicked.connect(apply_date_range)
        cancel_button.clicked.connect(dialog.reject)

        dialog.exec_()

    def apply_date_range(self, dialog):
        start_date = self.start_calendar.selectedDate().toPyDate()
        end_date = self.end_calendar.selectedDate().toPyDate()
        self.plot_graphs(start_date=start_date, end_date=end_date)
        dialog.accept()

    def set_light_theme(self):
        self.voltage_graph_widget.setBackground('w')
        self.current_graph_widget.setBackground('w')
        self.energy_graph_widget.setBackground('w')

        self.voltage_graph_widget.plot([], pen=pg.mkPen(color='b', width=2))  # Синя лінія для напруги
        self.current_graph_widget.plot([], pen=pg.mkPen(color='r', width=2))  # Червона лінія для струму
        self.energy_graph_widget.plot([], pen=pg.mkPen(color='g', width=2))  # Зелена лінія для енергії

    def open_export_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Експорт в Excel")
        dialog.setFixedSize(400, 150)
        layout = QVBoxLayout(dialog)

        # Вибір діапазону дат
        date_range_layout = QHBoxLayout()
        start_label = QLabel("Початок:")
        self.start_date = QDateEdit((QDate.currentDate().addYears(-1)))
        self.start_date.setCalendarPopup(True)
        end_label = QLabel("Кінець:")
        self.end_date = QDateEdit(QDate.currentDate())
        self.end_date.setCalendarPopup(True)
        date_range_layout.addWidget(start_label)
        date_range_layout.addWidget(self.start_date)
        date_range_layout.addWidget(end_label)
        date_range_layout.addWidget(self.end_date)

        layout.addLayout(date_range_layout)

        # Опції для включення середнього арифметичного та графіків
        self.include_average = QCheckBox("Додати середнє арифметичне")
        self.include_charts = QCheckBox("Додати графіки")
        self.include_average.setChecked(True)
        self.include_charts.setChecked(True)

        layout.addWidget(self.include_average)
        layout.addWidget(self.include_charts)

        # Кнопка для експорту
        save_button = QPushButton("Зберегти в Excel")
        save_button.clicked.connect(self.export_to_excel)

        layout.addWidget(save_button)

        dialog.setLayout(layout)
        dialog.exec()

    def export_to_excel(self):
        # Отримуємо початкову та кінцеву дату
        start_datetime = self.start_date.dateTime().toPyDateTime()
        end_datetime = self.end_date.dateTime().toPyDateTime()

        # Фільтруємо дані за діапазоном часу
        report_data = self.db_session.query(SDM120Report).filter(
            SDM120Report.device_id == self.device.id,
            SDM120Report.timestamp >= start_datetime,
            SDM120Report.timestamp <= end_datetime
        ).order_by(SDM120Report.timestamp).all()

        if not report_data:
            QMessageBox.warning(self, "Експорт", "Дані за вибраний період відсутні.")
            return

        # Перетворюємо дані у DataFrame
        data = [{
            "Час": report.timestamp,
            "Напруга (V)": report.voltage,
            "Струм (A)": report.current,
            "Енергія (kWh)": report.total_active_energy,
        } for report in report_data]

        df = pd.DataFrame(data)

        # Діалог для вибору шляху збереження
        save_path, _ = QFileDialog.getSaveFileName(self, "Зберегти файл", "", "Excel Files (*.xlsx);;All Files (*)")

        if save_path:
            wb = Workbook()
            ws = wb.active
            ws.title = "Дані"

            # Додаємо заголовки стовпців
            headers = ["Час", "Напруга (V)", "Струм (A)", "Енергія (kWh)"]
            ws.append(headers)

            # Якщо включено додавання середнього арифметичного
            if self.include_average.isChecked():
                # Додаємо дані до Excel з третього рядка
                ws.append(['', '', '', ''])
                for row in data:
                    ws.append([row['Час'], row['Напруга (V)'], row['Струм (A)'], row['Енергія (kWh)']])
                # Додаємо середнє значення для напруги та струму (в рядок 2)
                ws.cell(row=2, column=1,value="Середнє значення:")
                for col_idx, col in enumerate(df.columns[1:3], start=2):  # Лише для напруги та струму
                    col_letter = ws.cell(row=2, column=col_idx).column_letter
                    avg_formula = f"=AVERAGE({col_letter}3:{col_letter}{len(df) + 1})"
                    ws.cell(row=2, column=col_idx, value=avg_formula)
            else:
                for row in data:
                    ws.append([row['Час'], row['Напруга (V)'], row['Струм (A)'], row['Енергія (kWh)']])

            # Форматування заголовків
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Встановлюємо ширину стовпців у 120 пікселів
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 20  # Відповідає 120 пікселям

            if self.include_charts.isChecked():
                    # Додаємо графіки на окремий аркуш
                    charts_sheet = wb.create_sheet(title="Графіки")

                    # Графік напруги (V) - Лінійний графік
                    voltage_chart = LineChart()
                    voltage_chart.title = "Напруга (V)"
                    voltage_chart.style = 13
                    voltage_chart.x_axis.title = "Час"
                    voltage_chart.y_axis.title = "Напруга (V)"
                    voltage_chart.width = 20
                    voltage_chart.height = 10
                    voltage_chart.add_data(
                        Reference(ws, min_col=2, min_row=1, max_row=len(df) + 1), titles_from_data=True
                    )
                    voltage_chart.set_categories(
                        Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)  # Час для осі X
                    )
                    charts_sheet.add_chart(voltage_chart, "A1")

                    # Графік струму (A) - Лінійний графік
                    current_chart = LineChart()
                    current_chart.title = "Струм (A)"
                    current_chart.style = 13
                    current_chart.x_axis.title = "Час"
                    current_chart.y_axis.title = "Струм (A)"
                    current_chart.width = 20
                    current_chart.height = 10
                    current_chart.add_data(
                        Reference(ws, min_col=3, min_row=1, max_row=len(df) + 1), titles_from_data=True
                    )
                    current_chart.set_categories(
                        Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)  # Час для осі X
                    )
                    charts_sheet.add_chart(current_chart, "A20")

                    # Графік енергії (kWh) - Стовпчастий графік
                    energy_chart = BarChart()
                    energy_chart.title = "Енергія (kWh)"
                    energy_chart.style = 10
                    energy_chart.x_axis.title = "Час"
                    energy_chart.y_axis.title = "Енергія (kWh)"
                    energy_chart.width = 20
                    energy_chart.height = 10
                    energy_chart.add_data(
                        Reference(ws, min_col=4, min_row=1, max_row=len(df) + 1), titles_from_data=True
                    )
                    energy_chart.set_categories(
                        Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)  # Час для осі X
                    )
                    charts_sheet.add_chart(energy_chart, "A40")

            # Зберігаємо файл
            wb.save(save_path)
            QMessageBox.information(self, "Експорт", f"Дані успішно збережено в {save_path}")